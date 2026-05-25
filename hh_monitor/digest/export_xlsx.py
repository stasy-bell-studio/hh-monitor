"""Excel export for the digest candidate panel.

Produces an .xlsx workbook with:
  - 19 columns covering candidate identity, scores, LLM dossier (5 new), and status
  - Auto-filter on the header row
  - Frozen first row (freeze panes at A2)
  - Conditional fill on Score Total: green >= 70, yellow >= 60, red < 60
  - Data-validation dropdown on the Status column (matching screening_status values)
  - Hyperlinks in the URL column
  - «Комментарий LLM» column: populated only as fallback (when dossier fields absent)
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore[import-untyped]

from hh_monitor.digest.query import CandidateRow

# ── Column definitions ────────────────────────────────────────────────────────
# (header, width)
# Columns 1-14 mirror the original layout; 15-19 are the new dossier columns.
_COLUMNS: list[tuple[str, int]] = [
    ("#", 4),
    ("ID", 10),
    ("URL", 42),
    ("Текущая должность", 30),
    ("Регион", 18),
    ("Возраст", 8),
    ("Опыт (мес)", 10),
    ("Score Total", 11),
    ("Fit Score", 10),
    ("LLM Score", 10),
    ("Вердикт", 12),
    ("Комментарий LLM", 35),  # col 12 — fallback only when dossier absent
    ("Стоп-сигналы", 25),  # col 13 — fallback red_flags_str
    ("Статус", 16),  # col 14
    # ── Dossier (commit 9.3+) ──────────────────────────────────────────────
    ("Факты", 45),  # col 15
    ("Слабые места", 40),  # col 16
    ("Red flags", 35),  # col 17
    ("Вопросы интервью", 40),  # col 18
    ("Вердикт HR", 45),  # col 19
]

# Status dropdown column number (1-indexed): "Статус" is column 14
_STATUS_COL = 14
_STATUS_OPTIONS = ["taken", "reserve", "rejected", "contacts_opened"]

# Fill colours for Score Total
_FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
_FONT_LINK = Font(color="0563C1", underline="single", size=9)
_FONT_DATA = Font(size=9)
_ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="top")


def export_xlsx(candidates: list[CandidateRow], output_path: Path) -> None:
    """Write *candidates* to an Excel workbook at *output_path*.

    The file is created or overwritten.  The parent directory must exist.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Кандидаты"

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, (header, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Data rows ─────────────────────────────────────────────────────────────
    for rank, cand in enumerate(candidates, start=1):
        row_idx = rank + 1  # data starts at row 2

        # «Комментарий LLM» is only shown when dossier fields are absent (fallback)
        comment_cell_value = "" if cand.has_dossier else (cand.llm_comment or "")[:500]
        # «Стоп-сигналы» fallback — shows old JSONB red_flags list when no dossier
        red_flags_cell_value = "" if cand.has_dossier else cand.red_flags_str[:200]

        values: list[Any] = [
            rank,  # 1  #
            cand.hh_resume_id[:8],  # 2  ID
            cand.url,  # 3  URL
            cand.current_role[:60] if cand.current_role else "",  # 4 Текущая должность
            cand.region,  # 5  Регион
            cand.age,  # 6  Возраст
            cand.total_exp_months,  # 7  Опыт (мес)
            cand.score_total,  # 8  Score Total
            cand.fit_score,  # 9  Fit Score
            cand.llm_score,  # 10 LLM Score
            cand.llm_verdict or "",  # 11 Вердикт
            comment_cell_value,  # 12 Комментарий LLM (fallback)
            red_flags_cell_value,  # 13 Стоп-сигналы (fallback)
            cand.screening_status or "",  # 14 Статус
            (cand.dossier_facts_confirmed or "")[:800],  # 15 Факты
            (cand.dossier_weak_spots or "")[:600],  # 16 Слабые места
            (cand.dossier_red_flags or "")[:400],  # 17 Red flags
            cand.interview_questions_str[:400],  # 18 Вопросы интервью
            (cand.dossier_verdict or "")[:600],  # 19 Вердикт HR
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = _ALIGN_WRAP
            cell.font = _FONT_DATA

        # Hyperlink on URL cell (column 3)
        url_cell = ws.cell(row=row_idx, column=3)
        url_cell.hyperlink = cand.url
        url_cell.font = _FONT_LINK

        # Conditional fill on Score Total (column 8)
        score = cand.score_total
        score_cell = ws.cell(row=row_idx, column=8)
        if score is not None:
            if score >= 70:
                score_cell.fill = _FILL_GREEN
            elif score >= 60:
                score_cell.fill = _FILL_YELLOW
            else:
                score_cell.fill = _FILL_RED

    # ── Auto-filter on header row ─────────────────────────────────────────────
    last_col_letter = get_column_letter(len(_COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col_letter}1"

    # ── Freeze panes (keep header visible while scrolling) ───────────────────
    ws.freeze_panes = "A2"

    # ── Status column dropdown (column 14) ───────────────────────────────────
    if candidates:
        last_data_row = len(candidates) + 1
        status_col_letter = get_column_letter(_STATUS_COL)
        status_range = f"{status_col_letter}2:{status_col_letter}{last_data_row}"
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(_STATUS_OPTIONS)}"',
            allow_blank=True,
            showErrorMessage=False,
        )
        ws.add_data_validation(dv)
        dv.add(status_range)

    # ── Row height for data rows ──────────────────────────────────────────────
    ws.row_dimensions[1].height = 16
    for i in range(2, len(candidates) + 2):
        ws.row_dimensions[i].height = 50

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
