"""Styled multi-sheet Excel workbook for the weekly HR digest.

Each candidate is a unique person (one row per ``hh_resume_id``) with trend columns,
and every lifetime event for this week's candidates is surfaced on a dedicated history
sheet — the duplicates ARE the candidate's history, not noise to discard.

Sheet order:
  - «Сводка»            — cover: week/dates + one row per vacancy.
  - one sheet per vacancy — that position's full slice (one row per resume, score desc).
  - «Все кандидаты»     — the union of all vacancy sheets.
  - «По позициям»       — per-position funnel breakdown.
  - «Воронка»           — KPI cells + BarChart over the funnel stages.
  - «Динамика»          — weekly_series rows + LineChart.
  - «История кандидатов» — one row per all-time event for this week's candidates.

Style only is borrowed from hh_monitor/digest/export_xlsx.py (header fill/font, freeze
panes, auto-filter, widths, wrap). The «Статус скрининга» column here is a read-only
label — no DataValidation dropdown, no taken/reserve taxonomy.
"""

from __future__ import annotations

import ast
from io import BytesIO
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.chart import BarChart, LineChart, Reference  # type: ignore[import-untyped]
from openpyxl.formatting.rule import (  # type: ignore[import-untyped]
    ColorScaleRule,
    FormulaRule,
)
from openpyxl.styles import (  # type: ignore[import-untyped]
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from hh_monitor.weekly_digest.run import _status_label

if TYPE_CHECKING:
    from datetime import datetime

    from hh_monitor.weekly_digest.run import (
        _Candidate,
        _DigestData,
        _WeekPoint,
    )

_FILL_HEADER = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
_FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
_FONT_LINK = Font(color="0563C1", underline="single", size=9)
_FONT_DATA = Font(size=9)
_FONT_TITLE = Font(bold=True, size=14, color="1A3C6E")
_FONT_SUBTITLE = Font(size=10, color="595959")
_FONT_TOTAL = Font(bold=True, size=9)
_ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="top")

# Zebra striping + conditional-format fills.
_FILL_ZEBRA = PatternFill("solid", fgColor="F2F5FA")
_FILL_AMBER = PatternFill("solid", fgColor="FFF2CC")  # вердикт спорно
_FILL_GRAY = PatternFill("solid", fgColor="E7E6E6")  # вердикт мимо
_FILL_RED = PatternFill("solid", fgColor="F8CBAD")  # стоп-сигнал / отклонён
_FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")  # одобрен
_FILL_GROUP = PatternFill("solid", fgColor="DDEBF7")  # история: per-resume band

# История: Δ-direction font colors for the «Оценка» cell.
_FONT_UP = Font(size=9, bold=True, color="2E7D32")
_FONT_DOWN = Font(size=9, bold=True, color="C0392B")
_BORDER_GROUP_TOP = Border(top=Side(style="medium", color="1A3C6E"))

# Excel forbids these in a sheet name and caps it at 31 chars.
_FORBIDDEN_SHEET_CHARS = set(r":\/?*[]")
# Fixed sheet names reserved so a vacancy never collides with them.
_RESERVED_SHEETS = (
    "Сводка",
    "Все кандидаты",
    "По позициям",
    "Воронка",
    "Динамика",
    "История кандидатов",
)

# Candidate-sheet columns (shared by «Все кандидаты» and every vacancy sheet).
_CAND_HEADERS: list[tuple[str, int]] = [
    ("Позиция", 24),  # 1  A
    ("Оценка (тек.)", 11),  # 2  B  score_total (color scale here)
    ("Оценка (первая)", 13),  # 3  C  score_first
    ("Δ", 7),  # 4  D  score_delta
    ("Изменений", 10),  # 5  E  change_count
    ("Что менялось", 22),  # 6  F  change_types
    ("Соответствие портрету", 12),  # 7  G  fit_score
    ("Оценка ИИ", 10),  # 8  H  llm_score
    ("Вердикт", 14),  # 9  I  (conditional formatting)
    ("Регион", 18),  # 10 J
    ("Реальная роль", 22),  # 11 K
    ("Сильные стороны", 40),  # 12 L
    ("Слабые места", 36),  # 13 M
    ("Риски", 32),  # 14 N
    ("Вывод", 44),  # 15 O
    ("Статус скрининга", 16),  # 16 P  (conditional formatting)
    ("Причина", 30),  # 17 Q
    ("Ссылка", 14),  # 18 R  «Открыть →» hyperlink
    ("Дата", 12),  # 19 S
]
_CAND_SCORE_COL = 2
_CAND_VERDICT_COL = 9
_CAND_STATUS_COL = 16
_CAND_LINK_COL = 18

_EVENT_LABEL: dict[str, str] = {
    "NEW": "Появление",
    "REACTIVATED": "Возобновление",
    "REMOVED": "Снятие",
    "UPDATED_POSITION": "Должность",
    "UPDATED_SALARY": "Зарплата",
    "UPDATED_EXPERIENCE": "Опыт",
}


def _dict_lines(d: dict[object, object]) -> str:
    return "\n".join(f"{k}: {v}" for k, v in d.items())


def _humanize_field(value: object) -> str:
    """Render dict-valued LLM fields as readable `key: value` lines.

    The facts/weak/risks columns are mixed-type in the data: some rows hold plain
    text, others a ``dict`` or a single-quoted Python-dict repr string (e.g.
    ``"{'Опыт': '250+ агентов'}"``). Flatten any dict shape to newline-joined
    ``key: value`` lines; plain strings pass through unchanged.
    """
    if not value:
        return ""
    if isinstance(value, dict):
        return _dict_lines(value)
    if isinstance(value, str):
        if value.strip().startswith("{"):
            try:
                parsed = ast.literal_eval(value)
            except (ValueError, SyntaxError):
                return value  # malformed dict-ish text — keep as-is
            if isinstance(parsed, dict):
                return _dict_lines(parsed)
        return value
    return str(value)


def _sanitize_sheet_name(name: str, used: set[str]) -> str:
    """Return an Excel-safe, ≤31-char, collision-free sheet name.

    Strips forbidden ``: \\ / ? * [ ]`` chars, collapses whitespace, truncates to 31,
    and resolves collisions (case-insensitive, incl. reserved fixed sheets) with a
    `` (2)``, `` (3)`` … suffix kept within the 31-char budget.
    """
    cleaned = "".join(" " if ch in _FORBIDDEN_SHEET_CHARS else ch for ch in name)
    cleaned = " ".join(cleaned.split()).strip() or "Без названия"
    base = cleaned[:31]
    candidate = base
    n = 2
    while candidate.casefold() in used:
        suffix = f" ({n})"
        candidate = base[: 31 - len(suffix)].rstrip() + suffix
        n += 1
    used.add(candidate.casefold())
    return candidate


def _style_header(ws: Any, headers: list[tuple[str, int]]) -> None:
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    last = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last}1"


def _add_text_rule(ws: Any, col: int, n: int, needle: str, fill: PatternFill) -> None:
    """Add a ``ISNUMBER(SEARCH(needle, cell))`` conditional format over a data column.

    SEARCH is case-insensitive and substring-based, so it is robust to the emoji
    labels in «Статус скрининга» (e.g. matches "Одобрен" inside "Одобрен ✅").
    """
    if n <= 0:
        return
    letter = get_column_letter(col)
    rng = f"{letter}2:{letter}{n + 1}"
    formula = f'ISNUMBER(SEARCH("{needle}",{letter}2))'
    ws.conditional_formatting.add(rng, FormulaRule(formula=[formula], fill=fill))


def _write_candidate_sheet(ws: Any, candidates: list[_Candidate]) -> None:
    """Render one row per resume (already deduped) with trend columns + polish.

    Sorted upstream by score_total DESC. Freezes column A + header (B2), autofilter,
    zebra striping, a color scale on «Оценка (тек.)», conditional formatting on
    «Вердикт»/«Статус скрининга», and an «Открыть →» hyperlink in «Ссылка».
    """
    _style_header(ws, _CAND_HEADERS)
    ws.freeze_panes = "B2"  # freeze column A (Позиция) + header row

    for i, c in enumerate(candidates, start=2):
        delta = c["score_delta"]
        values = [
            c["position_name"],
            c["score_total"],
            c["score_first"],
            "" if delta is None else delta,
            c["change_count"],
            c["change_types"],
            c["fit_score"],
            c["llm_score"],
            c["llm_verdict"] or "",
            c["region"],
            c["llm_real_role"],
            _humanize_field(c["facts"]),
            _humanize_field(c["weak"]),
            _humanize_field(c["risks"]),
            c["conclusion"],
            _status_label(c["screening_status"]),
            c["reason"],
            "",  # «Ссылка» — set below as «Открыть →»
            c["created_at"].strftime("%d.%m.%Y"),
        ]
        zebra = i % 2 == 0
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col_idx, value=value)
            cell.alignment = _ALIGN_WRAP
            cell.font = _FONT_DATA
            if zebra:
                cell.fill = _FILL_ZEBRA
        if c["url"]:
            link_cell = ws.cell(row=i, column=_CAND_LINK_COL, value="Открыть →")
            link_cell.hyperlink = c["url"]
            link_cell.font = _FONT_LINK

    n = len(candidates)
    if not n:
        return
    score_letter = get_column_letter(_CAND_SCORE_COL)
    ws.conditional_formatting.add(
        f"{score_letter}2:{score_letter}{n + 1}",
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="F8696B",
            mid_type="num",
            mid_value=50,
            mid_color="FFEB84",
            end_type="num",
            end_value=100,
            end_color="63BE7B",
        ),
    )
    _add_text_rule(ws, _CAND_VERDICT_COL, n, "спорно", _FILL_AMBER)
    _add_text_rule(ws, _CAND_VERDICT_COL, n, "мимо", _FILL_GRAY)
    _add_text_rule(ws, _CAND_VERDICT_COL, n, "стоп", _FILL_RED)
    _add_text_rule(ws, _CAND_STATUS_COL, n, "Одобрен", _FILL_GREEN)
    _add_text_rule(ws, _CAND_STATUS_COL, n, "Отклонён", _FILL_RED)
    _add_text_rule(ws, _CAND_STATUS_COL, n, "Стоп-лист", _FILL_RED)


def _sheet_summary(
    wb: Workbook,
    data: _DigestData,
    week_num: int,
    date_from: datetime,
    date_to: datetime,
) -> None:
    """Cover sheet: week/dates title + one row per vacancy (+ a totals row)."""
    ws = wb.active
    ws.title = "Сводка"
    ws.cell(row=1, column=1, value=f"Еженедельная сводка · неделя {week_num}").font = _FONT_TITLE
    ws.cell(
        row=2,
        column=1,
        value=f"{date_from:%d.%m.%Y} – {date_to:%d.%m.%Y}",
    ).font = _FONT_SUBTITLE

    headers: list[tuple[str, int]] = [
        ("Позиция", 32),
        ("Найдено", 10),
        ("Спорно", 10),
        ("На ревью", 10),
        ("Отправлено", 11),
        ("Одобрено", 10),
    ]
    header_row = 4
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    pos_by_name = {pp["position_name"]: pp for pp in data["per_position"]}
    totals = {"found": 0, "doubt": 0, "pending": 0, "sent": 0, "approved": 0}
    row = header_row + 1
    for name in data["vacancies"]:
        pp = pos_by_name.get(name)
        if pp is None:
            continue
        vals = [name, pp["count"], pp["n_doubt"], pp["pending"], pp["sent"], pp["approved"]]
        for col_idx, value in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = _FONT_DATA
            cell.alignment = _ALIGN_WRAP
        totals["found"] += pp["count"]
        totals["doubt"] += pp["n_doubt"]
        totals["pending"] += pp["pending"]
        totals["sent"] += pp["sent"]
        totals["approved"] += pp["approved"]
        row += 1

    if data["vacancies"]:
        total_vals = [
            "Итого",
            totals["found"],
            totals["doubt"],
            totals["pending"],
            totals["sent"],
            totals["approved"],
        ]
        for col_idx, value in enumerate(total_vals, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = _FONT_TOTAL
    ws.freeze_panes = f"A{header_row + 1}"


def _sheet_positions(wb: Workbook, data: _DigestData) -> None:
    ws = wb.create_sheet("По позициям")
    headers: list[tuple[str, int]] = [
        ("Позиция", 24),
        ("Найдено", 9),
        ("Подходит", 9),
        ("Спорно", 9),
        ("Мимо", 9),
        ("Ср. рейтинг", 11),
        ("Отправлено", 11),
        ("Одобрено", 10),
        ("Отклонено", 10),
    ]
    _style_header(ws, headers)
    for i, pp in enumerate(data["per_position"], start=2):
        values = [
            pp["position_name"],
            pp["count"],
            pp["n_fit"],
            pp["n_doubt"],
            pp["n_miss"],
            pp["avg_score"],
            pp["sent"],
            pp["approved"],
            pp["rejected"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col_idx, value=value)
            cell.alignment = _ALIGN_WRAP
            cell.font = _FONT_DATA


def _sheet_funnel(wb: Workbook, data: _DigestData) -> None:
    ws = wb.create_sheet("Воронка")
    f = data["funnel"]
    conv = round(f["approved"] / f["sent"] * 100) if f["sent"] else 0
    ws.cell(row=1, column=1, value="Этап").font = _FONT_HEADER
    ws.cell(row=1, column=1).fill = _FILL_HEADER
    ws.cell(row=1, column=2, value="Кол-во").font = _FONT_HEADER
    ws.cell(row=1, column=2).fill = _FILL_HEADER
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 10
    stages = [
        ("Найдено", f["found"]),
        ("Отправлено", f["sent"]),
        ("Одобрено", f["approved"]),
        ("Отклонено", f["rejected"]),
    ]
    for i, (label, val) in enumerate(stages, start=2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val)
    extra_row = len(stages) + 2
    ws.cell(row=extra_row, column=1, value="Спорно")
    ws.cell(row=extra_row, column=2, value=f["doubt"])
    ws.cell(row=extra_row + 1, column=1, value="Ждут")
    ws.cell(row=extra_row + 1, column=2, value=f["pending"])
    ws.cell(row=extra_row + 2, column=1, value="Конверсия, %")
    ws.cell(row=extra_row + 2, column=2, value=conv)

    chart = BarChart()
    chart.title = "Воронка недели"
    chart.type = "col"
    chart.legend = None
    chart_data = Reference(ws, min_col=2, min_row=2, max_row=len(stages) + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(stages) + 1)
    chart.add_data(chart_data, titles_from_data=False)
    chart.set_categories(cats)
    ws.add_chart(chart, "D2")


def _sheet_dynamics(wb: Workbook, weekly_series: list[_WeekPoint]) -> None:
    ws = wb.create_sheet("Динамика")
    headers: list[tuple[str, int]] = [
        ("Неделя", 16),
        ("Найдено", 10),
        ("Отправлено", 11),
        ("Одобрено", 10),
    ]
    _style_header(ws, headers)
    for i, wp in enumerate(weekly_series, start=2):
        ws.cell(row=i, column=1, value=wp["week_label"])
        ws.cell(row=i, column=2, value=wp["found"])
        ws.cell(row=i, column=3, value=wp["sent"])
        ws.cell(row=i, column=4, value=wp["approved"])

    if weekly_series:
        last = len(weekly_series) + 1
        chart = LineChart()
        chart.title = "Динамика за 4 недели"
        chart_data = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=last)
        cats = Reference(ws, min_col=1, min_row=2, max_row=last)
        chart.add_data(chart_data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "F2")


def _sheet_history(wb: Workbook, data: _DigestData) -> None:
    """One row per lifetime event for this week's candidates, chronological per resume.

    Per-resume groups are visually distinct (alternating band + top border). The «Оценка»
    cell is colored by Δ direction vs the resume's previous *scored* event — None scores
    are skipped (empty cell, no color, no baseline reset), so a removal never reads as a
    score drop.
    """
    ws = wb.create_sheet("История кандидатов")
    headers: list[tuple[str, int]] = [
        ("Ссылка", 12),
        ("Дата", 12),
        ("Тип изменения", 18),
        ("Что менялось", 44),
        ("Оценка", 9),
        ("Вердикт", 14),
    ]
    _style_header(ws, headers)
    ws.freeze_panes = "B2"

    prev_rid: str | None = None
    prev_score: int | None = None
    group_idx = -1
    for row, h in enumerate(data["history"], start=2):
        rid = h["hh_resume_id"]
        new_group = rid != prev_rid
        if new_group:
            group_idx += 1
            prev_score = None  # new resume → reset Δ baseline
        band = _FILL_GROUP if group_idx % 2 == 1 else None

        score = h["score_total"]
        values = [
            "",  # «Ссылка» — set below
            h["created_at"].strftime("%d.%m.%Y"),
            _EVENT_LABEL.get(h["event_type"], h["event_type"]),
            h["change_desc"],
            "" if score is None else score,
            h["verdict"] or "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = _FONT_DATA
            cell.alignment = _ALIGN_WRAP
            if band is not None:
                cell.fill = band
            if new_group:
                cell.border = _BORDER_GROUP_TOP

        if h["url"]:
            link_cell = ws.cell(row=row, column=1, value="Открыть →")
            link_cell.hyperlink = h["url"]
            link_cell.font = _FONT_LINK

        if score is not None:
            score_cell = ws.cell(row=row, column=5)
            if prev_score is not None and score > prev_score:
                score_cell.font = _FONT_UP
            elif prev_score is not None and score < prev_score:
                score_cell.font = _FONT_DOWN
            prev_score = score  # baseline advances only on scored events
        prev_rid = rid


def build_digest_workbook(
    data: _DigestData,
    weekly_series: list[_WeekPoint],
    week_num: int,
    date_from: datetime,
    date_to: datetime,
) -> bytes:
    """Build the weekly-digest workbook and return its bytes.

    Robust to an empty week: with no candidates/vacancies/history it still produces a
    valid workbook (cover «Сводка» + the empty fixed sheets), never raising.
    """
    wb = Workbook()
    _sheet_summary(wb, data, week_num, date_from, date_to)

    cands_by_pos: dict[str, list[_Candidate]] = {}
    for c in data["candidates_all"]:
        cands_by_pos.setdefault(c["position_name"], []).append(c)

    used = {name.casefold() for name in _RESERVED_SHEETS}
    for name in data["vacancies"]:
        ws = wb.create_sheet(_sanitize_sheet_name(name, used))
        _write_candidate_sheet(ws, cands_by_pos.get(name, []))

    _write_candidate_sheet(wb.create_sheet("Все кандидаты"), data["candidates_all"])
    _sheet_positions(wb, data)
    _sheet_funnel(wb, data)
    _sheet_dynamics(wb, weekly_series)
    _sheet_history(wb, data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
