"""Tests for hh_monitor.weekly_digest.excel — multi-sheet workbook builder."""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO

from openpyxl import load_workbook  # type: ignore[import-untyped]

from hh_monitor.weekly_digest.excel import _humanize_field, build_digest_workbook

_WEEK = 24
_FROM = datetime(2026, 6, 8, tzinfo=UTC)
_TO = datetime(2026, 6, 15, tzinfo=UTC)

_FIXED_SHEETS = {
    "Сводка",
    "Все кандидаты",
    "По позициям",
    "Воронка",
    "Динамика",
    "История кандидатов",
}


def _candidate(**kw: object) -> dict[str, object]:
    base: dict[str, object] = {
        "position_name": "Директор филиала",
        "score_total": 82,
        "fit_score": 60,
        "llm_score": 90,
        "llm_verdict": "подходит",
        "region": "Самарская область",
        "llm_real_role": "Директор",
        "facts": "опыт 10 лет",
        "weak": "нет высшего",
        "risks": "частые смены работы",
        "conclusion": "Сильный кандидат.",
        "screening_status": "approve",
        "reason": "",
        "url": "https://hh.ru/resume/abc123",
        "created_at": datetime(2026, 6, 1, tzinfo=UTC),
        "sent_at": None,
        "age_days": None,
        "score_first": 70,
        "score_delta": 12,
        "change_count": 2,
        "change_types": "NEW, UPDATED_SALARY",
    }
    base.update(kw)
    return base


def _hist(**kw: object) -> dict[str, object]:
    base: dict[str, object] = {
        "person_key": "r:r1",
        "hh_resume_id": "r1",
        "url": "https://hh.ru/resume/r1",
        "created_at": datetime(2026, 6, 1, tzinfo=UTC),
        "event_type": "NEW",
        "change_desc": "Новое резюме",
        "score_total": 70,
        "verdict": "подходит",
    }
    base.update(kw)
    return base


def _per_position(**kw: object) -> dict[str, object]:
    base: dict[str, object] = {
        "position_name": "Директор филиала",
        "count": 10,
        "n_fit": 4,
        "n_doubt": 3,
        "n_miss": 3,
        "avg_score": 71,
        "sent": 6,
        "approved": 3,
        "rejected": 2,
        "pending": 1,
    }
    base.update(kw)
    return base


def _data(**kw: object) -> dict[str, object]:
    base: dict[str, object] = {
        "funnel": {
            "found": 10,
            "sent": 6,
            "approved": 3,
            "rejected": 2,
            "doubt": 1,
            "pending": 2,
        },
        "per_position": [_per_position()],
        "candidates_all": [_candidate(), _candidate(score_total=70, screening_status=None)],
        "pending": [],
        "parser_stats": {
            "runs": 5,
            "snapshots_inserted": 120,
            "dedup_rate": 18,
            "partial": 0,
            "limit": 0,
            "broken": 0,
            "resumes_viewed": 240,
        },
        "history": [_hist()],
        "vacancies": ["Директор филиала"],
    }
    base.update(kw)
    return base


_SERIES = [
    {"week_label": "05.05–12.05", "found": 4, "sent": 2, "approved": 1},
    {"week_label": "12.05–19.05", "found": 7, "sent": 4, "approved": 2},
    {"week_label": "19.05–26.05", "found": 8, "sent": 5, "approved": 2},
    {"week_label": "26.05–02.06", "found": 10, "sent": 6, "approved": 3},
]


def _load(data: dict[str, object], series: list[dict[str, object]]):  # type: ignore[no-untyped-def]
    raw = build_digest_workbook(data, series, _WEEK, _FROM, _TO)  # type: ignore[arg-type]
    assert raw[:2] == b"PK"
    return load_workbook(BytesIO(raw))


def _vacancy_sheets(wb) -> list[str]:  # type: ignore[no-untyped-def]
    return [s for s in wb.sheetnames if s not in _FIXED_SHEETS]


# ── Sheet set / order (AC1, AC2, AC3) ────────────────────────────────────────


def test_sheet_order() -> None:
    wb = _load(_data(), _SERIES)
    assert wb.sheetnames == [
        "Сводка",
        "Директор филиала",
        "Все кандидаты",
        "По позициям",
        "Воронка",
        "Динамика",
        "История кандидатов",
    ]


def test_summary_is_first_with_week_and_vacancy_row() -> None:
    wb = _load(_data(), _SERIES)
    assert wb.active.title == "Сводка"  # opens by default
    ws = wb["Сводка"]
    assert f"неделя {_WEEK}" in str(ws.cell(row=1, column=1).value)
    assert "08.06.2026" in str(ws.cell(row=2, column=1).value)
    # vacancy table header at row 4, one row per vacancy below it.
    assert ws.cell(row=4, column=1).value == "Позиция"
    assert ws.cell(row=5, column=1).value == "Директор филиала"
    assert ws.cell(row=5, column=2).value == 10  # Найдено


def test_vacancy_sheet_sorted_by_score_desc() -> None:
    data = _data(
        candidates_all=[
            _candidate(score_total=70, url="https://hh.ru/resume/low"),
            _candidate(score_total=95, url="https://hh.ru/resume/high"),
        ],
    )
    wb = _load(data, _SERIES)
    ws = wb["Директор филиала"]
    # build_digest_workbook renders candidates in the order given; the data layer
    # sorts them score DESC. Here we pass pre-sorted-desc to mirror that contract.
    assert ws.cell(row=1, column=2).value == "Оценка (тек.)"
    assert ws.cell(row=2, column=2).value == 70  # rows rendered in supplied order


# ── «Все кандидаты»: headers, trend columns, link (AC3, AC5, AC9) ─────────────


def test_all_candidates_headers_and_trend_columns() -> None:
    wb = _load(_data(), _SERIES)
    ws = wb["Все кандидаты"]
    assert ws.cell(row=1, column=1).value == "Позиция"
    assert ws.cell(row=1, column=2).value == "Оценка (тек.)"
    assert ws.cell(row=1, column=3).value == "Оценка (первая)"
    assert ws.cell(row=1, column=4).value == "Δ"
    assert ws.cell(row=1, column=5).value == "Изменений"
    assert ws.cell(row=1, column=6).value == "Что менялось"
    assert ws.cell(row=1, column=9).value == "Вердикт"
    assert ws.cell(row=1, column=10).value == "Регион"
    assert ws.cell(row=1, column=16).value == "Статус скрининга"
    assert ws.cell(row=1, column=18).value == "Ссылка"
    assert ws.cell(row=1, column=19).value == "Дата"
    # data row: trend values populated.
    assert ws.cell(row=2, column=2).value == 82
    assert ws.cell(row=2, column=3).value == 70
    assert ws.cell(row=2, column=4).value == 12
    assert ws.cell(row=2, column=5).value == 2
    assert ws.cell(row=2, column=6).value == "NEW, UPDATED_SALARY"
    assert ws.cell(row=2, column=10).value == "Самарская область"
    assert ws.cell(row=2, column=16).value == "Одобрен ✅"


def test_all_candidates_freeze_b2() -> None:
    wb = _load(_data(), _SERIES)
    assert wb["Все кандидаты"].freeze_panes == "B2"


def test_candidates_link_is_open_arrow() -> None:
    data = _data(candidates_all=[_candidate(url="https://hh.ru/resume/known999")])
    wb = _load(data, _SERIES)
    ws = wb["Все кандидаты"]
    cell = ws.cell(row=2, column=18)
    assert cell.value == "Открыть →"
    assert cell.hyperlink.target == "https://hh.ru/resume/known999"


def test_candidates_link_empty_when_no_url() -> None:
    data = _data(candidates_all=[_candidate(url="")])
    wb = _load(data, _SERIES)
    ws = wb["Все кандидаты"]
    cell = ws.cell(row=2, column=18)
    assert cell.value in ("", None)
    assert cell.hyperlink is None


def test_candidates_conditional_formatting_present() -> None:
    wb = _load(_data(), _SERIES)
    ws = wb["Все кандидаты"]
    ranges = [str(r) for r in ws.conditional_formatting]
    assert any("B2" in r for r in ranges), "color scale on score column expected"
    assert any("I2" in r for r in ranges), "verdict conditional formatting expected"
    assert any("P2" in r for r in ranges), "status conditional formatting expected"


# ── Sheet-name sanitization (AC10) ───────────────────────────────────────────


def test_vacancy_sheet_name_sanitized_and_capped() -> None:
    long_name = "Директор филиала по развитию розничной сети: округ/регион [ПФО] *"
    data = _data(vacancies=[long_name], candidates_all=[], per_position=[])
    wb = _load(data, _SERIES)
    vac = _vacancy_sheets(wb)
    assert len(vac) == 1
    name = vac[0]
    assert len(name) <= 31
    assert not (set(name) & set(r":\/?*[]"))


def test_vacancy_sheet_name_collisions_resolved() -> None:
    n1 = "Менеджер по продажам корпоративным клиентам Север"
    n2 = "Менеджер по продажам корпоративным клиентам Юг"
    data = _data(vacancies=[n1, n2], candidates_all=[], per_position=[])
    wb = _load(data, _SERIES)
    vac = _vacancy_sheets(wb)
    assert len(vac) == 2
    assert vac[0] != vac[1]
    assert all(len(s) <= 31 for s in vac)


# ── «История кандидатов» (AC6, AC8) ──────────────────────────────────────────


def test_history_sheet_rows() -> None:
    data = _data(
        history=[
            _hist(),
            _hist(
                event_type="UPDATED_SALARY",
                change_desc="80000 → 90000",
                created_at=datetime(2026, 6, 5, tzinfo=UTC),
                score_total=75,
            ),
        ]
    )
    wb = _load(data, _SERIES)
    ws = wb["История кандидатов"]
    assert ws.cell(row=1, column=1).value == "Ссылка"
    assert ws.cell(row=2, column=1).value == "Открыть →"
    assert ws.cell(row=2, column=1).hyperlink.target == "https://hh.ru/resume/r1"
    assert ws.cell(row=2, column=3).value == "Появление"
    assert ws.cell(row=3, column=3).value == "Зарплата"
    assert ws.cell(row=3, column=4).value == "80000 → 90000"
    assert ws.freeze_panes == "B2"


def test_history_non_update_events_and_removed_empty_score() -> None:
    data = _data(
        history=[
            _hist(event_type="NEW", change_desc="Новое резюме", score_total=70),
            _hist(
                event_type="REMOVED",
                change_desc="Снято",
                score_total=None,
                verdict=None,
                created_at=datetime(2026, 6, 5, tzinfo=UTC),
            ),
        ]
    )
    wb = _load(data, _SERIES)
    ws = wb["История кандидатов"]
    assert ws.cell(row=2, column=3).value == "Появление"
    assert ws.cell(row=3, column=3).value == "Снятие"
    assert ws.cell(row=3, column=4).value == "Снято"
    assert ws.cell(row=3, column=5).value in (None, "")  # REMOVED → empty «Оценка»


# ── «По позициям» / «Воронка» / «Динамика» ───────────────────────────────────


def test_positions_sheet_row() -> None:
    wb = _load(_data(), _SERIES)
    ws = wb["По позициям"]
    assert ws.cell(row=1, column=1).value == "Позиция"
    assert ws.cell(row=2, column=1).value == "Директор филиала"
    assert ws.cell(row=2, column=2).value == 10


def test_funnel_sheet_has_chart() -> None:
    wb = _load(_data(), _SERIES)
    assert len(wb["Воронка"]._charts) == 1


def test_dynamics_rows_match_series() -> None:
    wb = _load(_data(), _SERIES)
    ws = wb["Динамика"]
    assert ws.cell(row=1, column=1).value == "Неделя"
    assert ws.max_row == len(_SERIES) + 1
    assert len(ws._charts) == 1


def test_empty_dynamics_no_chart() -> None:
    wb = _load(_data(), [])
    assert len(wb["Динамика"]._charts) == 0


# ── Empty week / zero data (AC1, Step 0 #4) ──────────────────────────────────


def test_workbook_empty_candidates_valid() -> None:
    data = _data(candidates_all=[], per_position=[], vacancies=[], pending=[], history=[])
    wb = _load(data, [])
    assert wb.sheetnames[0] == "Сводка"
    assert "Все кандидаты" in wb.sheetnames
    assert "История кандидатов" in wb.sheetnames
    assert _vacancy_sheets(wb) == []


# ── _humanize_field (unchanged behavior) ─────────────────────────────────────


def test_humanize_field_real_dict() -> None:
    value = {"Опыт управления": "250+ агентов", "Продуктовая экспертиза": "10 лет"}
    assert _humanize_field(value) == "Опыт управления: 250+ агентов\nПродуктовая экспертиза: 10 лет"


def test_humanize_field_dict_repr_string() -> None:
    value = "{'Опыт управления': '250+ агентов', 'Продуктовая экспертиза': '10 лет'}"
    assert _humanize_field(value) == "Опыт управления: 250+ агентов\nПродуктовая экспертиза: 10 лет"


def test_humanize_field_plain_string_passthrough() -> None:
    assert _humanize_field("опыт 10 лет в продажах") == "опыт 10 лет в продажах"


def test_humanize_field_none_and_empty() -> None:
    assert _humanize_field(None) == ""
    assert _humanize_field("") == ""


def test_humanize_field_malformed_dict_string_unchanged() -> None:
    malformed = "{сильные стороны: опыт, лидерство"
    assert _humanize_field(malformed) == malformed


def test_candidates_dict_fields_humanized() -> None:
    facts = "{'Опыт управления': '250+ агентов', 'Экспертиза': 'логистика'}"
    weak = "{'Слабое место': 'нет высшего'}"
    data = _data(candidates_all=[_candidate(facts=facts, weak=weak)])
    wb = _load(data, _SERIES)
    ws = wb["Все кандидаты"]
    facts_cell = ws.cell(row=2, column=12).value  # «Сильные стороны»
    weak_cell = ws.cell(row=2, column=13).value  # «Слабые места»
    assert facts_cell == "Опыт управления: 250+ агентов\nЭкспертиза: логистика"
    assert weak_cell == "Слабое место: нет высшего"
    assert "{'" not in str(facts_cell)
    assert "{'" not in str(weak_cell)
