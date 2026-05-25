"""Tests for hh_monitor.digest — query, xlsx export, pdf export.

Coverage for commits 9.2 and 9.3:

  test_fetch_candidates_returns_matching_rows
    — Seeds Search + Resume + Snapshot + Event, calls fetch_candidates,
      asserts the seeded candidate appears in results.

  test_export_xlsx_creates_valid_workbook
    — Builds CandidateRow objects (with dossier fields), calls export_xlsx,
      asserts the output has correct 19-column header + data rows.

  test_export_xlsx_fallback_llm_comment
    — CandidateRow without dossier → «Комментарий LLM» cell populated from
      llm_comment; dossier columns empty.

  test_export_pdf_creates_non_empty_file
    — Creates PDF; checks file exists and is non-empty (skipped if WeasyPrint
      system libs not available).

  test_export_pdf_dossier_sections
    — PDF HTML contains dossier section heading markers.
"""

from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any

import pytest
from sqlalchemy.ext.asyncio import AsyncSession

from hh_monitor.db.models import Event, Resume, Search, Snapshot
from hh_monitor.digest.query import CandidateRow, fetch_candidates

# ── Helpers ───────────────────────────────────────────────────────────────────


def _hash(payload: dict[str, Any]) -> str:
    return hashlib.sha256(json.dumps(payload, sort_keys=True).encode()).hexdigest()


def _make_candidate(
    rid: str = "abc00000000000001",
    score_total: int = 75,
    llm_verdict: str = "подходит",
    *,
    with_dossier: bool = False,
) -> CandidateRow:
    """Build an in-memory CandidateRow for unit tests."""
    dossier_kwargs: dict[str, Any] = {}
    if with_dossier:
        dossier_kwargs = {
            "dossier_facts_confirmed": "Кандидат работал в СОГАЗ 4 года (2019–2023).",
            "dossier_weak_spots": "Нет данных о P&L. Gap с 2023 не объяснён.",
            "dossier_red_flags": "Короткие сроки (<1.5 года) в двух предыдущих компаниях.",
            "dossier_interview_questions": [
                "Каков был реальный размер вашей агентской сети?",
                "Чем занимались с 2023 по 2024?",
            ],
            "dossier_verdict": "Рекомендую на первый звонок.",
        }

    return CandidateRow(
        hh_resume_id=rid,
        score_total=score_total,
        fit_score=60,
        llm_score=80,
        llm_verdict=llm_verdict,
        llm_comment="Хороший кандидат",
        llm_red_flags=["Нет опыта ОСАГО"],
        screening_status=None,
        payload={
            "id": rid,
            "title": "Директор филиала",
            "age": 38,
            "area": {"id": "2", "name": "Санкт-Петербург"},
            "total_experience": {"months": 96},
            "experience": [
                {
                    "company": "ВСК Страхование",
                    "position": "Директор филиала",
                    "start": "2019-01",
                    "end": None,
                }
            ],
        },
        **dossier_kwargs,
    )


# ── AC4a: query ───────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_fetch_candidates_returns_matching_rows(db_session: AsyncSession) -> None:
    """fetch_candidates returns the seeded candidate for the given search_code."""
    sc = Search(
        search_code="digest_test_sc",
        position_code="branch_director",
        position_name="Директор филиала",
        hh_params={"text": "директор"},
        portrait={"position_code": "branch_director", "position_name": "Директор филиала"},
    )
    db_session.add(sc)
    await db_session.flush()

    rid = "digest0000000001"
    db_session.add(Resume(hh_resume_id=rid, score_total=70, fit_score=55, llm_score=78))
    await db_session.flush()

    payload: dict[str, Any] = {
        "id": rid,
        "title": "Директор филиала",
        "total_experience": {"months": 60},
    }
    db_session.add(Snapshot(hh_resume_id=rid, payload=payload, content_hash=_hash(payload)))
    await db_session.flush()

    db_session.add(
        Event(
            hh_resume_id=rid,
            event_type="NEW",
            search_id=sc.id,
            details={"curr_snapshot_id": 1},
        )
    )
    await db_session.flush()

    candidates = await fetch_candidates(
        db_session,
        search_code="digest_test_sc",
        min_score=60,
        include_screened=False,
    )

    assert len(candidates) >= 1, "Expected at least one candidate row"
    ids = [c.hh_resume_id for c in candidates]
    assert rid in ids, f"Seeded resume {rid!r} missing from results: {ids}"


@pytest.mark.asyncio
async def test_fetch_candidates_respects_min_score(db_session: AsyncSession) -> None:
    """fetch_candidates excludes resumes below min_score."""
    sc = Search(
        search_code="digest_test_minscore",
        position_code="branch_director",
        position_name="Директор филиала",
        hh_params={"text": "директор"},
        portrait={"position_code": "branch_director", "position_name": "Директор филиала"},
    )
    db_session.add(sc)
    await db_session.flush()

    rid = "digest0000000002"
    db_session.add(Resume(hh_resume_id=rid, score_total=45))  # below threshold
    await db_session.flush()

    payload: dict[str, Any] = {"id": rid, "title": "Менеджер"}
    db_session.add(Snapshot(hh_resume_id=rid, payload=payload, content_hash=_hash(payload)))
    await db_session.flush()

    db_session.add(Event(hh_resume_id=rid, event_type="NEW", search_id=sc.id, details={}))
    await db_session.flush()

    candidates = await fetch_candidates(
        db_session, search_code="digest_test_minscore", min_score=60
    )
    assert all(c.hh_resume_id != rid for c in candidates), "Resume below min_score must not appear"


# ── AC4b: xlsx export ─────────────────────────────────────────────────────────


def test_export_xlsx_creates_valid_workbook(tmp_path: Path) -> None:
    """export_xlsx creates .xlsx with 19-column header and correct rank values."""
    from openpyxl import load_workbook

    from hh_monitor.digest.export_xlsx import _COLUMNS, export_xlsx

    # Use candidates with dossier fields to exercise new columns
    candidates = [
        _make_candidate("r0000000000000001", 75, with_dossier=True),
        _make_candidate("r0000000000000002", 65, with_dossier=True),
    ]
    out = tmp_path / "test_digest.xlsx"

    export_xlsx(candidates, out)

    assert out.exists(), "xlsx file was not created"
    wb = load_workbook(str(out))
    ws = wb.active

    # Header row: 19 columns
    assert len(_COLUMNS) == 19, f"Expected 19 columns, got {len(_COLUMNS)}"
    headers = [ws.cell(row=1, column=i + 1).value for i in range(len(_COLUMNS))]
    expected_headers = [col[0] for col in _COLUMNS]
    assert headers == expected_headers, f"Headers mismatch: {headers}"

    # Two data rows, rank column (col 1) = 1 and 2
    data_rows = [ws.cell(row=r, column=1).value for r in range(2, len(candidates) + 2)]
    assert len(data_rows) == 2, f"Expected 2 data rows, got {len(data_rows)}"
    assert data_rows[0] == 1
    assert data_rows[1] == 2

    # Dossier columns (15–19) should be populated
    col_facts = ws.cell(row=2, column=15).value  # «Факты»
    assert col_facts and "СОГАЗ" in col_facts, f"Expected facts text, got: {col_facts!r}"

    col_verdict = ws.cell(row=2, column=19).value  # «Вердикт HR»
    assert col_verdict and "Рекомендую" in col_verdict


def test_export_xlsx_fallback_llm_comment(tmp_path: Path) -> None:
    """When no dossier fields, xlsx uses llm_comment in col 12 (fallback)."""
    from openpyxl import load_workbook

    from hh_monitor.digest.export_xlsx import export_xlsx

    # Candidate without dossier (legacy enriched record)
    candidates = [_make_candidate("r_legacy_001", 70, with_dossier=False)]
    out = tmp_path / "test_fallback.xlsx"
    export_xlsx(candidates, out)

    wb = load_workbook(str(out))
    ws = wb.active

    # «Комментарий LLM» is column 12 — should contain llm_comment
    comment_cell = ws.cell(row=2, column=12).value
    assert comment_cell == "Хороший кандидат", f"Expected fallback comment, got: {comment_cell!r}"

    # Dossier columns 15–19 should be empty
    for col in range(15, 20):
        assert ws.cell(row=2, column=col).value in (None, ""), (
            f"Expected empty dossier col {col}, got: {ws.cell(row=2, column=col).value!r}"
        )


# ── AC4c: pdf export ──────────────────────────────────────────────────────────


def test_export_pdf_creates_non_empty_file(tmp_path: Path) -> None:
    """export_pdf creates a non-empty .pdf file (skipped if WeasyPrint not available)."""
    try:
        import weasyprint as _wp  # noqa: F401
    except OSError:
        pytest.skip("WeasyPrint system libs (pango/cairo) not available")

    from hh_monitor.digest.export_pdf import export_pdf

    candidates = [_make_candidate()]
    out = tmp_path / "test_digest.pdf"

    export_pdf(candidates, out, search_code="branch_director_21vek")

    assert out.exists(), "pdf file was not created"
    assert out.stat().st_size > 1024, (
        f"pdf file too small ({out.stat().st_size} bytes) — likely empty"
    )


def test_export_pdf_dossier_sections(tmp_path: Path) -> None:
    """When dossier fields present, rendered HTML contains dossier section markers."""
    from jinja2 import Environment, FileSystemLoader, select_autoescape

    from hh_monitor.digest.export_pdf import _TEMPLATES_DIR

    env = Environment(
        loader=FileSystemLoader(str(_TEMPLATES_DIR)),
        autoescape=select_autoescape(["html"]),
    )
    template = env.get_template("card.html")

    # Candidate with full dossier
    candidate_with_dossier = _make_candidate("r_dos_001", 80, with_dossier=True)
    html = template.render(candidates=[candidate_with_dossier], search_code="test")

    assert 'class="dossier-facts"' in html, "Expected .dossier-facts div in HTML"
    assert 'class="dossier-weak"' in html, "Expected .dossier-weak div in HTML"
    assert 'class="dossier-flags"' in html, "Expected .dossier-flags div in HTML"
    assert 'class="dossier-verdict"' in html, "Expected .dossier-verdict div in HTML"
    # Dossier questions rendered as <li> items
    assert 'class="dossier-questions"' in html or "Агентской сети" in html


def test_export_pdf_fallback_to_llm_comment(tmp_path: Path) -> None:
    """Without dossier fields, rendered HTML shows legacy llm_comment block."""
    from jinja2 import Environment, FileSystemLoader, select_autoescape

    from hh_monitor.digest.export_pdf import _TEMPLATES_DIR

    env = Environment(
        loader=FileSystemLoader(str(_TEMPLATES_DIR)),
        autoescape=select_autoescape(["html"]),
    )
    template = env.get_template("card.html")

    # Candidate WITHOUT dossier (legacy path)
    legacy_candidate = _make_candidate("r_leg_001", 70, with_dossier=False)
    html = template.render(candidates=[legacy_candidate], search_code="test")

    assert "comment-box" in html, "Expected .comment-box in fallback HTML"
    assert "Хороший кандидат" in html, "Expected llm_comment text in fallback"
    # Should NOT render dossier-specific section *elements* (CSS class defs are always present)
    assert 'class="dossier-facts"' not in html
